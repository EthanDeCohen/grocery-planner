"""Create the tracked Excel template workbook with sheets preloaded from sample CSVs."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
TEMPLATE = ROOT / "template" / "GroceryPlanner.template.xlsx"

STORES = [
    ("Whole Foods", "wholefoods", "Whole Foods", "Whole Foods Deals"),
    ("Food Lion", "foodlion", "Food Lion", "Food Lion Deals"),
    ("Harris Teeter", "harristeeter", "Harris Teeter", "Harris Teeter Deals"),
]

HEADER_FILL = PatternFill("solid", start_color="E6F0FA")
HEADER_FONT = Font(bold=True, name="Arial")


def load_csv(path: Path, store: str, row_type: str) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    df = pd.read_csv(path)
    if df.empty:
        return df
    df.insert(0, "row_type", row_type)
    df.insert(0, "store", store)
    return df


def write_sheet(ws, df: pd.DataFrame, placeholder: str) -> None:
    ws.delete_rows(1, ws.max_row or 1)
    if df.empty:
        ws["A1"] = placeholder
        return

    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)


def build() -> Path:
    TEMPLATE.parent.mkdir(parents=True, exist_ok=True)

    all_prices = []
    all_deals = []
    sheets: dict[str, pd.DataFrame | str] = {}

    for store_name, folder, prices_sheet, deals_sheet in STORES:
        prices = load_csv(DATA / folder / "prices.csv", store_name, "price")
        deals = load_csv(DATA / folder / "deals.csv", store_name, "deal")
        sheets[prices_sheet] = prices
        sheets[deals_sheet] = deals
        if not prices.empty:
            all_prices.append(prices)
        if not deals.empty:
            all_deals.append(deals)

    sheets["All Prices"] = pd.concat(all_prices, ignore_index=True) if all_prices else pd.DataFrame()
    sheets["All Deals"] = pd.concat(all_deals, ignore_index=True) if all_deals else pd.DataFrame()
    sheets["Savings Summary"] = pd.DataFrame(
        {
            "Metric": ["Total price rows", "Total deal rows", "Data folder"],
            "Value": [
                len(sheets["All Prices"]) if isinstance(sheets["All Prices"], pd.DataFrame) else 0,
                len(sheets["All Deals"]) if isinstance(sheets["All Deals"], pd.DataFrame) else 0,
                str(DATA),
            ],
        }
    )

    instructions = pd.DataFrame(
        {
            "Step": [
                "Grocery Planner",
                "",
                "Last refreshed",
                "Refresh timestamp",
                "Refresh summary",
                "",
                "How to use",
                "1",
                "2",
                "3",
                "4",
                "",
                "Git",
            ],
            "Detail": [
                "",
                "",
                "(run RefreshGroceryData macro after importing VBA)",
                "",
                "",
                "",
                "",
                "Copy template/GroceryPlanner.template.xlsx to GroceryPlanner.xlsm in the project root.",
                "Import vba/*.cls and vba/*.bas into the workbook (or run scripts/import_vba.ps1).",
                "Update CSV files under data/<store>/prices.csv and deals.csv.",
                "Run RefreshGroceryData (Alt+F8) to reload all store sheets.",
                "",
                "Tracked: template/, vba/, scripts/. Gitignored: data/, GroceryPlanner.xlsm.",
            ],
        }
    )

    with pd.ExcelWriter(TEMPLATE, engine="openpyxl") as writer:
        instructions.to_excel(writer, sheet_name="Instructions", index=False)
        for name, content in sheets.items():
            if isinstance(content, pd.DataFrame):
                content.to_excel(writer, sheet_name=name, index=False)
            else:
                pd.DataFrame({"info": [content]}).to_excel(writer, sheet_name=name, index=False)

    wb = load_workbook(TEMPLATE)
    ws = wb["Instructions"]
    ws["A1"].font = Font(bold=True, size=16, name="Arial")
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for sheet_name in wb.sheetnames:
        if sheet_name == "Instructions":
            continue
        write_sheet(wb[sheet_name], sheets.get(sheet_name, pd.DataFrame()), f"Run RefreshGroceryData for {sheet_name}")
    wb.save(TEMPLATE)
    return TEMPLATE


if __name__ == "__main__":
    print(f"Created {build()}")